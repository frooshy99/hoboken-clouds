#! python3
# a spending log program.

from time import sleep
import sys
from datetime import date
from datetime import datetime
import openpyxl
import numpy as np
import pandas as pd


def p(text, delay=0.005, end=''):
    for c in text:
        sys.stdout.write(c)
        sys.stdout.flush()
        sleep(delay)


def pi(text): print(text, end='')


def exi(inp):
    if str(inp).casefold() == 'exit' or str(inp).casefold() == 'e':
        p('Goodbye.\n')
        sys.exit()


def input_(argv_ind):
    if len(sys.argv) < argv_ind + 1:
        text = input()

        exi(text)
        return text
    else:
        text = sys.argv[argv_ind]
        text = str(text).replace('_', ' ')  # use underscores in place of spaces for command line input

        p(text + '\n')
        return text


# add any shortcuts you want
shortcuts = {'b': 'breakfast',
             'l': 'lunch',
             'd': 'dinner',
             'g': 'groceries'}


monthDic = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June',
            7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'}


def space(string): return ' ' * len(string)


def stars(num): return '*' * num


# this updates time of last entry to the nearest second.
def update():
    f = open('logtime.txt', 'w')
    entrytime = str(datetime.now())[:19]

    f.write(entrytime)
    f.close()


wbDir = 'logbook.xlsx'
wb = openpyxl.load_workbook(wbDir)
ws = wb.active

# reads last entry time
ent = open('logtime.txt')
p(f'Last entry: {ent.read()}')

# feel free to change this to whatever
cats = ['debit card', 'credit card']

p('\nEnter \'exit\' at any prompt to stop program.\n\n')

while True:
    # checks number of filled rows.
    for x in range(1, 10000):
        if ws.cell(row=x, column=1).value is None:
            numFilled = int(x)

            break
        else:
            continue

    # to create an empty dataframe
    index = np.arange(numFilled)
    columns = ['YYYY', 'MM', 'DD', 'EXPENSE', 'SPENT', 'CATEGORY']
    data = np.array([np.arange(numFilled)] * 6).T

    df = pd.DataFrame(index=index, columns=columns, data=data)

    dateL = []
    nameL = []
    spentL = []
    catL = []

    for i in range(1, numFilled):
        dateL.append(str(ws[f'A{i}'].value)[:10])
        nameL.append(str(ws[f'B{i}'].value))
        spentL.append(str(ws[f'C{i}'].value))
        catL.append(str(ws[f'D{i}'].value))

    def conv(to_conv): return pd.Series(to_conv)

    dateL = conv(dateL)
    nameL = conv(nameL)
    spentL = conv(spentL)
    catL = conv(catL)

    dateDF = dateL.str.split('-', expand=True)

    # populating the df
    df['YYYY'] = dateDF[0]
    df['MM'] = dateDF[1]
    df['DD'] = dateDF[2]
    df['EXPENSE'] = nameL
    df['SPENT'] = spentL
    df['CATEGORY'] = catL
    
    def to_float(text): return float(text)
    df['SPENT_fl'] = df['SPENT'].apply(to_float)

    p('1 - Add to log\n'
      '2 - View summary for month\n'
      'What would you like to do? ')
    choice = input_(1)

    # this will be for adding to log
    if int(choice) == 1:
        p('\nHow much did you spend? ')
        val = input_(2)

        p('What did you spend it on? ')
        name = input_(3)

        # for shortcuts
        if str(name) in shortcuts:
            name = shortcuts[name]

        for i in range(len(cats)):
            p(f'{i + 1} - {cats[i]}\n')
            
        p('Which category would you like to add this to? ')
        cat = input_(4)

        d = date.today()

        for x in range(1, 10000):
            if ws.cell(row=x, column=1).value is None:
                ws[f'A{x}'] = d
                ws[f'B{x}'] = name
                ws[f'C{x}'] = val
                ws[f'D{x}'] = cat
                break
            else:
                continue

        wb.save(wbDir)
        p('\nData added to log.\n')
        update()

        if len(sys.argv) > 1:
              sys.exit()

        continue

    elif int(choice) == 2:
        date_today = str(date.today())
        YYYY_now = date_today[:4]
        MM_now = date_today[5:7]

        df_month = df[(df['YYYY'] == YYYY_now) & (df['MM'] == MM_now)]

        # if still no data for current month
        if len(df_month) < 1:
            MM_now = str(int(MM_now) - 1)
            df_month = df[(df['YYYY'] == YYYY_now) & (df['MM'] == MM_now)]
        
        days = df_month['DD'].unique().tolist()

        p(f'\n{monthDic[int(MM_now)]} {YYYY_now}:\n')

        for i in range(len(days)):
            pi(f'{days[i]}: ')

            today = df_month[df_month['DD'] == days[i]]
            today = today[['EXPENSE', 'SPENT', 'CATEGORY']]

            for j in range(len(today)):
                info = today.iloc[j]
                expense = info['EXPENSE']
                spent = info['SPENT']
                category = int(info['CATEGORY'])

                if j > 0:
                    p(space(f'{days[i]}: '))

                pi(f'{expense} | ${spent} | {stars(category)}\n')

        spent_month = df_month['SPENT_fl'].sum()

        budget = 600
        budget_frac = int(spent_month / budget * 50)
        
        over_warning = ''
        if budget_frac > 50:
            budget_frac = 50
            over_warning = ' Overbudget!'
              
        bar = '█'
        spac = ' '
        under = f'Spent: ${round(spent_month, 2)}'
        
        # "progress" bar
        p(f'\n|{bar * budget_frac}{spac * (50 - budget_frac)}|{over_warning}\n')
        pi(under + ' ' * (37 - len(under)) + f'Budget: ${budget}.00\n')

        meals = list(shortcuts.values())[:3]
        others = list(shortcuts.values())[3:]
        # things not in meals & not in others = misc

        # categorical breakdown of expenses, only the meal expenses are grouped together
        breakdown = {}
                     
        def br_update(text, df_df):
            if df_df['SPENT_fl'].sum() > 0:  # problems to alignment (line 303) etc. if 0 value item gets into dic
                breakdown.update({text: float('%.2f' % df_df['SPENT_fl'].sum())})
        
        df_month_meals = df_month[df_month['EXPENSE'].isin(meals)]
        br_update('meals', df_month_meals)

        for i in range(len(others)):
            spentspent = df_month[df_month['EXPENSE'] == others[i]]
            br_update(others[i], spentspent)

        misc = df_month[~df_month['EXPENSE'].isin(meals) & ~df_month['EXPENSE'].isin(others)]
        br_update('miscellaneous', misc)

        breakdown = sorted(breakdown.items(), key=lambda a: a[1], reverse=True)
        # descending sort

        prev_pipe = 0

        for i in range(len(breakdown)):
            amount = breakdown[i][1]
            pipe = int(amount / budget * 50)

            if i == len(breakdown) - 1 and prev_pipe + pipe < budget_frac:
                pipe = budget_frac - prev_pipe
            # gotta do this to make sure things align
            # not most accurate presentation but yeah
            
            p(' ' * (prev_pipe + 1) + '|' * pipe + f' {breakdown[i][0]}, ${amount}\n')
            prev_pipe = pipe + prev_pipe
            
        for i in range(len(cats)):
            if i < 1:
                p(f'\nSpent on ')
            else:
                p(space('Spent on '))

            df_cat = df_month[df_month['CATEGORY'] == f'{i + 1}']
            p(f'{cats[i]}: ${round(df_cat["SPENT_fl"].sum(), 2)}\n')


        p('\n5 most costly this month:\n')
        top5 = df_month.nlargest(n=5, columns='SPENT_fl')

        for i in range(len(top5)):
            row = top5.iloc[i]
            exp = row['EXPENSE']
            amt = row['SPENT_fl']

            p(f'{i + 1}. {exp}, ${amt}\n')

        sys.exit()
